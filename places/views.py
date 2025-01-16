from datetime import datetime
from io import BytesIO

import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.workbook import Workbook
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import WeatherReport, Place


class ImportPlacesAPIView(APIView):
    """Ендпоинт для импорта xlsx-файла с данными"""
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Файл не предоставлен"}, status=400)

        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                name = row[0]
                location = row[1]
                latitude, longitude = location.split(",")
                rating = row[2]

                Place.objects.create(
                    name=name,
                    location=f"POINT({longitude} {latitude})",
                    rating=rating
                )

            return Response({"message": "Данные успешно импортированы"})
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class WeatherExportAPIView(APIView):
    """
    Ендпоинт для экспорта xlsx-файла с отчетом погоды за определенный период.

    Пример запроса:
    http://127.0.0.1:8000/export_weather/?place_id=1&start_date=2025-01-01&end_date=2025-01-16
    """

    def get(self, request, *args, **kwargs):
        place_id = request.query_params.get('place_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if not place_id or not start_date or not end_date:
            return Response({
                'error': 'Нехватает параметров запроса: ID места, дата начала замера, дата конца замера'
            }, status=400)

        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))

            end_date = datetime.strptime(end_date, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
        except ValueError:
            return Response({'error': 'Неверный формат дат. Ожидается формат YYYY-MM-DD.'}, status=400)

        file_name = f"weather_report_{place_id}_{start_date.date()}_{end_date.date()}"

        weather_data = WeatherReport.objects.filter(
            place__id=place_id,
            created_at__gte=start_date,
            created_at__lte=end_date
        )

        if not weather_data:
            return Response({'error': 'Нет данных для выбранного периода и места.'}, status=404)

        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = file_name

        base_headers = [field.name for field in WeatherReport._meta.fields]
        headers = [field.verbose_name for field in WeatherReport._meta.fields]
        ws.append(headers)

        for data in weather_data:
            row = []
            for field in base_headers:
                value = str(getattr(data, field))
                row.append(value)
            ws.append(row)

        wb.save(output)
        output.seek(0)

        response = HttpResponse(output.getvalue(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={ws.title}'

        return response
